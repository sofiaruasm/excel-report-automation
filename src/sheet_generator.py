import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
 
 
COR_CABECALHO = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
FONTE_CABECALHO = Font(color="FFFFFF", bold=True)
 
 
def gerar_relatorio(df: pd.DataFrame, caminho_saida: str) -> None:
    """
    Gera o relatório em Excel com:
      - Aba 'Resumo': receita total por categoria, com gráfico de barras
      - Aba 'Detalhado': todos os dados linha a linha
 
    Args:
        df: DataFrame já limpo (ver data_cleaner.py).
        caminho_saida: caminho onde o arquivo .xlsx final será salvo.
    """
    df = df.copy()
    df["receita"] = df["quantidade"] * df["preco_unitario"]
 
    wb = Workbook()
 
    _criar_aba_resumo(wb, df)
    _criar_aba_detalhado(wb, df)
 
    wb.save(caminho_saida)
    print(f"Relatório gerado com sucesso em: {caminho_saida}")
 
 
def _criar_aba_resumo(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.active
    ws.title = "Resumo"
 
    resumo = (
        df.groupby("categoria")["receita"]
        .sum()
        .reset_index()
        .sort_values("receita", ascending=False)
    )
 
    ws["A1"] = "Receita por Categoria"
    ws["A1"].font = Font(bold=True, size=14)
 
    linha_cabecalho = 3
    for col_idx, titulo in enumerate(["Categoria", "Receita Total (R$)"], start=1):
        celula = ws.cell(row=linha_cabecalho, column=col_idx, value=titulo)
        celula.font = FONTE_CABECALHO
        celula.fill = COR_CABECALHO
        celula.alignment = Alignment(horizontal="center")
 
    primeira_linha_dados = linha_cabecalho + 1
    for i, linha in enumerate(resumo.itertuples(index=False), start=primeira_linha_dados):
        ws.cell(row=i, column=1, value=linha.categoria)
        ws.cell(row=i, column=2, value=round(linha.receita, 2))
 
    _ajustar_largura_colunas(ws)
    _adicionar_grafico_barras(ws, linha_cabecalho, primeira_linha_dados, len(resumo))
 
 
def _criar_aba_detalhado(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Detalhado")
 
    colunas = ["data", "produto", "categoria", "regiao", "quantidade", "preco_unitario", "receita"]
    cabecalhos = ["Data", "Produto", "Categoria", "Região", "Qtd. Vendida", "Preço Unit. (R$)", "Receita (R$)"]
 
    for col_idx, titulo in enumerate(cabecalhos, start=1):
        celula = ws.cell(row=1, column=col_idx, value=titulo)
        celula.font = FONTE_CABECALHO
        celula.fill = COR_CABECALHO
 
    for row_idx, linha in enumerate(df[colunas].itertuples(index=False), start=2):
        ws.cell(row=row_idx, column=1, value=linha.data.strftime("%d/%m/%Y"))
        ws.cell(row=row_idx, column=2, value=linha.produto)
        ws.cell(row=row_idx, column=3, value=linha.categoria)
        ws.cell(row=row_idx, column=4, value=linha.regiao)
        ws.cell(row=row_idx, column=5, value=linha.quantidade)
        ws.cell(row=row_idx, column=6, value=linha.preco_unitario)
        ws.cell(row=row_idx, column=7, value=round(linha.receita, 2))
 
    _ajustar_largura_colunas(ws)
 
 
def _adicionar_grafico_barras(ws, linha_cabecalho: int, primeira_linha_dados: int, num_linhas: int) -> None:
    grafico = BarChart()
    grafico.title = "Receita por Categoria"
    grafico.y_axis.title = "Receita (R$)"
    grafico.x_axis.title = "Categoria"
    grafico.style = 10  
 
    ultima_linha_dados = primeira_linha_dados + num_linhas - 1
    dados = Reference(ws, min_col=2, min_row=linha_cabecalho, max_row=ultima_linha_dados)
    categorias = Reference(ws, min_col=1, min_row=primeira_linha_dados, max_row=ultima_linha_dados)
 
    grafico.add_data(dados, titles_from_data=True)
    grafico.set_categories(categorias)
    grafico.width = 16
    grafico.height = 9
 
    ws.add_chart(grafico, "D3")
 
 
def _ajustar_largura_colunas(ws, folga: int = 2) -> None:
    for coluna_celulas in ws.columns:
        maior_tamanho = max(
            len(str(celula.value)) if celula.value is not None else 0
            for celula in coluna_celulas
        )
        letra_coluna = get_column_letter(coluna_celulas[0].column)
        ws.column_dimensions[letra_coluna].width = maior_tamanho + folga
 
 
if __name__ == "__main__":
    from data_reader import ler_dados
    from data_cleaner import limpar_dados
 
    df_bruto = ler_dados("vendas_bagunçado.xlsx")
    df_limpo = limpar_dados(df_bruto)
    gerar_relatorio(df_limpo, "relatorio_final.xlsx")